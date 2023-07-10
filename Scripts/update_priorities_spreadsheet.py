import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def copy_worksheet(source, destination):
    destination.sheet_properties.tabColor = source.sheet_properties.tabColor
    for row in source.iter_rows():
        for cell in row:
            destination[cell.coordinate].value = cell.value
            destination[cell.coordinate].data_type = cell.data_type
            destination[cell.coordinate].hyperlink = cell.hyperlink
    for col_idx in range(1, source.max_column + 1):
        col_letter = get_column_letter(col_idx)
        destination.column_dimensions[col_letter].width = source.column_dimensions[col_letter].width
    for row_idx in range(1, source.max_row + 1):
        destination.row_dimensions[row_idx].height = source.row_dimensions[row_idx].height


def update_spreadsheet(file_path):
    # Define other variables
    new_sheet_name = datetime.now().strftime("%m-%d-%Y")  # Today's date as the new sheet name

    # Load the Excel file with openpyxl directly
    excel_file = load_workbook(file_path)

    # Check if a sheet with the same name exists, and delete it if found
    if new_sheet_name in excel_file.sheetnames:
        del excel_file[new_sheet_name]


    # Use the last sheet as sheet_name
    sheet_names = excel_file.sheetnames
    if len(sheet_names) > 0:
        sheet_name = sheet_names[-1]

    # Load the template sheet
    template_sheet = excel_file[sheet_name]

    # Create a new sheet and copy values and formatting from the template sheet
    new_sheet = excel_file.create_sheet(title=new_sheet_name)
    copy_worksheet(template_sheet, new_sheet)

    # Rename cell A1 with today's date
    new_sheet["A1"] = new_sheet_name

    # Update cell C2 with the next Monday after today's date
    next_monday = datetime.now() + timedelta(days=5)
    new_sheet["C2"] = next_monday.strftime("%m-%d-%Y")



    # Iterate through each row in the sheet
    for row in new_sheet.iter_rows(min_row=3, max_row=new_sheet.max_row):

        # Get the cell in column E for each row
        cell = row[4]  # Index 4 corresponds to column E (0-based indexing)

        # Check if the cell contains a string and the word "completed" (case-insensitive) is present or is None
        if isinstance(cell.value, str) and ("completed" in cell.value.lower() or cell.value is None):
            # Store the value of the corresponding row in column C ("Description")
            cell_b = row[2].value
            #write the value of cell_b to in the last row of column A
            new_sheet["A" + str(new_sheet.max_row + 1)] = f"Deleted {cell_b}"


            # Delete the row
            new_sheet.delete_rows(cell.row)

        else:
            # do nothing
            pass

    # Save the updated spreadsheet
    excel_file.save(file_path)

# Define the file path and sheet name
file_path = r"C:\Users\grellama\OneDrive - Continental Resources\Personel\weekly updates\project_tracker_update.xlsx"

# Call the update_spreadsheet function
update_spreadsheet(file_path)

# Open the updated Excel file

os.startfile(file_path)